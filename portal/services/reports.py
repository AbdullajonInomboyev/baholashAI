"""Zam dekan uchun kafedra/fakultet hisobotlarini Excel'ga chiqarish."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.db.models import Avg, Count

from teaching.models import ResourceReview, TeacherAssignment


HEAD_FILL = PatternFill("solid", fgColor="274690")
SUB_FILL = PatternFill("solid", fgColor="EFEAF7")


def _title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, size=14, color="16233B")
    cell.alignment = Alignment(horizontal="left")


def _header_row(ws, row, headers):
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        ws.column_dimensions[get_column_letter(col)].width = max(16, len(str(title)) + 3)


def department_report(department):
    """Kafedra kesimi: fanlar, o‘qituvchilar, resurs sifat ballari."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kafedra hisoboti"
    _title(ws, f"Kafedra hisoboti — {department.name}", 6)

    headers = ["Fan kodi", "Fan nomi", "O‘qituvchi(lar)", "Resurslar", "O‘rtacha moslik %", "O‘rtacha to‘liqlik %"]
    _header_row(ws, 3, headers)

    row = 4
    for link in department.course_links.select_related("course").order_by("course__code"):
        assignments = (
            TeacherAssignment.objects
            .filter(department_course=link, status=TeacherAssignment.Status.ACTIVE)
            .select_related("teacher")
        )
        teachers = ", ".join(a.teacher.full_name or a.teacher.username for a in assignments) or "—"
        reviews = ResourceReview.objects.filter(
            resource__links__teacher_assignment__department_course=link,
            status=ResourceReview.Status.COMPLETED,
        )
        agg = reviews.aggregate(n=Count("id"), m=Avg("match_score"), c=Avg("completeness_score"))
        ws.cell(row=row, column=1, value=link.course.code)
        ws.cell(row=row, column=2, value=link.course.name)
        ws.cell(row=row, column=3, value=teachers)
        ws.cell(row=row, column=4, value=agg["n"] or 0)
        ws.cell(row=row, column=5, value=round(agg["m"], 1) if agg["m"] else None)
        ws.cell(row=row, column=6, value=round(agg["c"], 1) if agg["c"] else None)
        row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="Bu kafedraga hali fan biriktirilmagan.")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def faculty_report(faculty):
    """Fakultet kesimi: kafedralar bo‘yicha yig‘ma ko‘rsatkichlar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fakultet hisoboti"
    _title(ws, f"Fakultet hisoboti — {faculty.name}", 5)

    headers = ["Kafedra", "Mudir", "Fanlar", "O‘qituvchilar", "O‘rtacha resurs sifati %"]
    _header_row(ws, 3, headers)

    row = 4
    for dept in faculty.departments.all().order_by("name"):
        course_count = dept.course_links.count()
        teacher_count = (
            TeacherAssignment.objects
            .filter(department_course__department=dept, status=TeacherAssignment.Status.ACTIVE)
            .values("teacher").distinct().count()
        )
        quality = ResourceReview.objects.filter(
            resource__links__teacher_assignment__department_course__department=dept,
            status=ResourceReview.Status.COMPLETED,
        ).aggregate(avg=Avg("match_score"))["avg"]
        ws.cell(row=row, column=1, value=dept.name)
        ws.cell(row=row, column=2, value=(dept.head.full_name if dept.head else "—"))
        ws.cell(row=row, column=3, value=course_count)
        ws.cell(row=row, column=4, value=teacher_count)
        ws.cell(row=row, column=5, value=round(quality, 1) if quality else None)
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
