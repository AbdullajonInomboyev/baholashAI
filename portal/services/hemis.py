from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.crypto import get_random_string
from django.utils.text import slugify

from accounts.models import PasswordStatus, Role, StudentProfile, UserRole

User = get_user_model()

# HEMIS eksporti sarlavhalari (shablon uchun, aynan shu tartibda)
HEMIS_HEADERS = [
    "Talaba ID", "To‘liq ismi", "Fuqarolik", "Davlat", "Millat", "Viloyat", "Tuman",
    "Jins", "Tug‘ilgan sana", "Pasport raqami", "JSHSHIR-kod", "Pasport berilgan sana",
    "Kurs", "Fakultet", "Guruh", "Ta'lim tili", "O‘quv yili", "Semestr", "Bitiruvchi",
    "Mutaxassislik", "Ta'lim turi", "Ta'lim shakli", "To‘lov shakli", "Grant turi",
    "Avvalgi ta'lim ma'lumoti", "Talaba toifasi", "Ijtimoiy toifa",
    "Birga yashaydiganlar soni", "Birga yashaydiganlar toifasi", "Yashash joyi statusi",
    "Yashash joyi geolokatsiyasi", "Buyruq", "GPA", "Kontrakt №", "Shartnoma turi",
]

# (sarlavha ichidagi kalit, maydon nomi, tur) — birinchi mos kelgan qoida ishlaydi.
# Kalitlar apostrofsiz (normalizatsiyadan keyin). Aniqroq qoidalar birinchi.
FIELD_RULES = [
    ("talaba id", "student_id", "str"),
    ("jshshir", "pinfl", "str"), ("pinfl", "pinfl", "str"),
    ("pasport raqami", "passport_number", "str"),
    ("pasport berilgan", "passport_issued", "date"),
    ("tug", "birth_date", "date_user"),
    ("jins", "gender", "gender"),
    ("fuqarolik", "citizenship", "str"),
    ("davlat", "country", "str"),
    ("millat", "nationality", "str"),
    ("viloyat", "region", "str"),
    ("tuman", "district", "str"),
    ("toliq ismi", "full_name", "name_user"), ("fish", "full_name", "name_user"),
    ("mutaxassislik", "specialty_code", "str"),
    ("fakultet", "faculty_name", "str"),
    ("guruh", "group_name", "str"),
    ("talim tili", "education_language", "str"), ("tili", "education_language", "str"),
    ("oquv yili", "academic_year", "str"),
    ("semestr", "semester", "str"),
    ("bitiruvchi", "is_graduating", "bool"),
    ("talim turi", "education_type", "str"),
    ("talim shakli", "education_form", "str"),
    ("tolov shakli", "payment_form", "str"),
    ("grant", "grant_type", "str"),
    ("avvalgi ta", "previous_education", "str"),
    ("talaba toifasi", "student_category", "str"),
    ("ijtimoiy", "social_category", "str"),
    ("yashaydiganlar soni", "cohabitants_count", "int"),
    ("yashaydiganlar toifasi", "cohabitants_category", "str"),
    ("yashash joyi statusi", "residence_status", "str"),
    ("geolokatsiya", "residence_geo", "str"),
    ("buyruq", "order_info", "str"),
    ("gpa", "gpa", "dec"),
    ("kontrakt", "contract_number", "str"),
    ("shartnoma", "contract_type", "str"),
    ("kurs", "course", "str"),
    ("ismi", "full_name", "name_user"),
]

_YES = {"ha", "yes", "1", "+", "true", "bor"}


def _norm(s):
    s = str(s or "").strip().lower()
    for ch in "‘’'`ʻʼ":
        s = s.replace(ch, "")
    return " ".join(s.split())


def _field_for(header):
    h = _norm(header)
    if not h:
        return None
    for key, field, kind in FIELD_RULES:
        if key in h:
            return field, kind
    return None


def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month"):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_gender(v):
    s = _norm(v)
    if s.startswith("erkak") or s in {"m", "male", "erkak"}:
        return StudentProfile.Gender.MALE
    if s.startswith("ayol") or s in {"f", "female", "ayol"}:
        return StudentProfile.Gender.FEMALE
    return ""


def _parse_dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).replace(",", ".").strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_int(v):
    if v in (None, ""):
        return None
    digits = "".join(ch for ch in str(v) if ch.isdigit())
    return int(digits) if digits else None


def _styled_header(ws, headers):
    fill = PatternFill("solid", fgColor="274690")
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(title) + 3)


def build_template():
    wb = Workbook(); ws = wb.active; ws.title = "Talabalar (HEMIS)"
    _styled_header(ws, HEMIS_HEADERS)
    ws.append([
        "456221100575", "XASANBOYEV XUMOYUN ABDUMAJID O‘G‘LI", "O‘zbekiston Respublikasi fuqarosi",
        "O‘zbekiston", "O‘zbeklar", "Namangan viloyati", "Yangiqo‘rg‘on tumani", "Erkak",
        "1999-06-17", "AB0181591", "31706995950017", "2015-06-19", "4-kurs",
        "Biznes boshqaruvi fakulteti", "ATDT-22-AU", "O‘zbek", "2025-2026", "8-semestr", "Ha",
        "60610800", "Bakalavr", "Kunduzgi", "To‘lov-shartnoma", "11 - Kontrakt",
        "2015-2018, Kollej", "Oddiy", "Boshqa", "4", "Kursdoshlari", "", "",
        "№ 01-T / 05.09.2022", "4.92", "", "",
    ])
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue()


def _unique_username(base):
    base = (base or "talaba").strip() or "talaba"
    cand = base; i = 1
    while User.objects.filter(username=cand).exists():
        i += 1; cand = f"{base}{i}"
    return cand


@transaction.atomic
def import_hemis(file_obj):
    """HEMIS Excel faylidan talabalarni yaratadi/yangilaydi (JSHSHIR yoki Talaba ID bo'yicha)."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {"created": 0, "updated": 0, "errors": ["Fayl bo‘sh."], "total": 0}

    # ustun indeksi -> (field, kind)
    colmap = {}
    for idx, header in enumerate(header_row):
        m = _field_for(header)
        if m:
            colmap[idx] = m

    if not any(f in {"pinfl", "student_id"} for f, _ in colmap.values()):
        return {"created": 0, "updated": 0,
                "errors": ["'JSHSHIR-kod' yoki 'Talaba ID' ustuni topilmadi."], "total": 0}

    created = updated = total = 0
    errors = []

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        total += 1

        prof_vals = {}
        user_full_name = None
        user_birth = None
        pinfl = student_id = ""

        for idx, (field, kind) in colmap.items():
            raw = row[idx] if idx < len(row) else None
            if kind == "name_user":
                if raw:
                    user_full_name = str(raw).strip()
            elif kind == "date_user":
                user_birth = _parse_date(raw)
            elif kind == "date":
                prof_vals[field] = _parse_date(raw)
            elif kind == "gender":
                prof_vals[field] = _parse_gender(raw)
            elif kind == "bool":
                prof_vals[field] = _norm(raw) in _YES
            elif kind == "int":
                prof_vals[field] = _parse_int(raw)
            elif kind == "dec":
                prof_vals[field] = _parse_dec(raw)
            else:
                prof_vals[field] = str(raw).strip() if raw is not None else ""

        pinfl = prof_vals.get("pinfl", "") or ""
        student_id = prof_vals.get("student_id", "") or ""

        if not pinfl and not student_id:
            errors.append(f"{ridx}-qator: JSHSHIR ham, Talaba ID ham bo‘sh — o‘tkazib yuborildi.")
            total -= 1
            continue

        # Mavjud talabani topamiz (avval JSHSHIR, keyin Talaba ID)
        profile = None
        if pinfl:
            profile = StudentProfile.objects.filter(pinfl=pinfl).select_related("user").first()
        if not profile and student_id:
            profile = StudentProfile.objects.filter(student_id=student_id).select_related("user").first()

        if profile:
            user = profile.user
            is_new = False
        else:
            base = student_id or pinfl or slugify(user_full_name or "talaba").replace("-", "_")
            user = User.objects.create(
                username=_unique_username(str(base)),
                full_name=user_full_name or "",
            )
            user.set_password(get_random_string(10))
            user.password_status = PasswordStatus.TEMPORARY
            user.save()
            profile = StudentProfile(user=user)
            is_new = True

        # User maydonlari
        changed = []
        if user_full_name and user.full_name != user_full_name:
            user.full_name = user_full_name; changed.append("full_name")
        if user_birth and user.birth_date != user_birth:
            user.birth_date = user_birth; changed.append("birth_date")
        if changed:
            user.save(update_fields=changed)

        UserRole.objects.get_or_create(user=user, role=Role.STUDENT)

        # Profil maydonlari
        for field, val in prof_vals.items():
            setattr(profile, field, val if val is not None else getattr(profile, field))
        # bo'sh bo'lsa ham to'g'ridan-to'g'ri yozamiz (None emas)
        for field, val in prof_vals.items():
            if val is None and field in {"passport_issued", "gpa", "cohabitants_count"}:
                setattr(profile, field, None)
        profile.save()

        created += 1 if is_new else 0
        updated += 0 if is_new else 1

    return {"created": created, "updated": updated, "errors": errors, "total": total}