"""Talabalarni Excel orqali import va export qilish."""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.crypto import get_random_string
from django.utils.text import slugify

from academics.models import AcademicYear, Direction, StudentEnrollment, StudyForm
from accounts.models import PasswordStatus, Role, UserRole

User = get_user_model()

HEADERS = ["F.I.Sh.", "Login", "Email", "Telefon", "Yo‘nalish kodi", "Ta‘lim shakli", "Guruh", "Holat",
           "Imkoniyati cheklangan", "Imkoniyat turi"]
FORM_LABELS = {label.lower(): value for value, label in StudyForm.choices}
STATUS_LABELS = {label.lower(): value for value, label in StudentEnrollment.Status.choices}
_YES = {"ha", "yes", "1", "+", "true", "bor", "x"}


def _disability_type(val):
    from accounts.models import User as _U
    if not val:
        return ""
    m = {label.lower(): value for value, label in _U.DisabilityType.choices if value}
    return m.get(str(val).strip().lower(), "")


def _styled_header(ws):
    fill = PatternFill("solid", fgColor="274690")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(title) + 4)


def build_template():
    """Bo‘sh, namunali import shablonini qaytaradi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    _styled_header(ws)
    ws.append(["Aliyev Vali", "aliyev_vali", "aliyev@namdu.uz", "+998901234567",
               "60610100", "Kunduzgi", "ATT-25-01", "O‘qimoqda", "yo‘q", ""])
    ws.append(["Karimova Dilnoza", "karimova_d", "", "+998907654321",
               "60610100", "Kunduzgi", "ATT-25-01", "O‘qimoqda", "ha", "Ko‘rish"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_students(academic_year: AcademicYear):
    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    _styled_header(ws)
    enrollments = (
        StudentEnrollment.objects
        .filter(academic_year=academic_year)
        .select_related("student", "direction")
        .order_by("group_name", "student__full_name")
    )
    for e in enrollments:
        ws.append([
            e.student.full_name, e.student.username, e.student.email, e.student.phone,
            e.direction.code, e.get_study_form_display(), e.group_name, e.get_status_display(),
            "ha" if e.student.is_disabled else "yo‘q", e.student.get_disability_type_display() if e.student.is_disabled else "",
        ])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@transaction.atomic
def import_students(academic_year: AcademicYear, file_obj):
    """Excel'dan talabalarni yaratadi/yangilaydi va o‘quv yiliga biriktiradi."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    faculty = academic_year.faculty
    directions = {d.code: d for d in Direction.objects.filter(faculty=faculty)}

    created, updated, errors = 0, 0, []
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        full_name, login, email, phone, dir_code, form, group, status, disabled, dis_type = (list(row) + [None] * 10)[:10]
        if not full_name or not dir_code:
            errors.append(f"{index}-qator: F.I.Sh. yoki yo‘nalish kodi bo‘sh.")
            continue
        direction = directions.get(str(dir_code).strip())
        if not direction:
            errors.append(f"{index}-qator: '{dir_code}' yo‘nalishi bu fakultetda topilmadi.")
            continue

        username = str(login).strip() if login else _make_login(str(full_name))
        is_disabled = bool(disabled) and str(disabled).strip().lower() in _YES
        dtype = _disability_type(dis_type) if is_disabled else ""
        user, is_new = User.objects.get_or_create(
            username=username,
            defaults={"full_name": str(full_name).strip(),
                      "email": (str(email).strip() if email else ""),
                      "phone": (str(phone).strip() if phone else "")},
        )
        user.is_disabled = is_disabled
        user.disability_type = dtype
        if is_disabled:
            user.tts_enabled = True
        if is_new:
            user.set_password(get_random_string(10))
            user.password_status = PasswordStatus.TEMPORARY
            user.save()
            created += 1
        else:
            user.full_name = str(full_name).strip()
            user.save(update_fields=["full_name", "is_disabled", "disability_type", "tts_enabled"])
            updated += 1

        UserRole.objects.get_or_create(user=user, role=Role.STUDENT)
        StudentEnrollment.objects.update_or_create(
            student=user, academic_year=academic_year,
            defaults={
                "direction": direction,
                "study_form": FORM_LABELS.get(str(form).strip().lower(), StudyForm.FULL_TIME) if form else StudyForm.FULL_TIME,
                "group_name": str(group).strip() if group else "",
                "status": STATUS_LABELS.get(str(status).strip().lower(), StudentEnrollment.Status.ACTIVE) if status else StudentEnrollment.Status.ACTIVE,
            },
        )
    return {"created": created, "updated": updated, "errors": errors}


def _make_login(full_name):
    base = slugify(full_name).replace("-", "_") or "talaba"
    candidate = base
    suffix = 1
    while User.objects.filter(username=candidate).exists():
        suffix += 1
        candidate = f"{base}{suffix}"
    return candidate
