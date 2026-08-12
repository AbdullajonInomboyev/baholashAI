from django.contrib import messages
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from academics.models import StudentEnrollment
from core.notifications import notify_many
from django.urls import reverse
from django.contrib.auth import get_user_model
from accounts.models import Role
from assessment.models import Choice, Question, Quiz, QuizAnswer, QuizAttempt
from assessment.services import quiz as quiz_service
from teaching.models import TeacherAssignment

from .access import role_required
from .forms_quiz import QuestionForm, QuizForm

STATUS = TeacherAssignment.Status


# ==================== O'QITUVCHI ====================

def _teacher_quizzes(user):
    return (Quiz.objects
            .filter(teacher_assignment__teacher=user)
            .select_related("teacher_assignment__department_course__course"))


@role_required(Role.TEACHER)
def teacher_quizzes(request):
    ctx = {"panel_title": "O‘qituvchi paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "quizzes": _teacher_quizzes(request.user),
           "form": QuizForm(teacher=request.user)}
    return render(request, "portal/quiz/teacher_list.html", ctx)


@role_required(Role.TEACHER)
def quiz_create(request):
    form = QuizForm(request.POST or None, teacher=request.user)
    if request.method == "POST" and form.is_valid():
        quiz = form.save()
        cur = quiz.teacher_assignment.department_course.course.curriculum
        students = get_user_model().objects.filter(
            enrollments__direction=cur.direction, enrollments__study_form=cur.study_form).distinct()
        notify_many(students, f"Yangi test: {quiz.title}", reverse("portal:student_quizzes"))
        messages.success(request, "Test yaratildi. Endi savol qo‘shing.")
        return redirect("portal:quiz_manage", pk=quiz.pk)
    return redirect("portal:teacher_quizzes")


@role_required(Role.TEACHER)
def quiz_manage(request, pk):
    quiz = get_object_or_404(_teacher_quizzes(request.user), pk=pk)
    ctx = {"panel_title": "O‘qituvchi paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "quiz": quiz,
           "questions": quiz.questions.prefetch_related("choices"),
           "form": QuestionForm()}
    return render(request, "portal/quiz/manage.html", ctx)


@role_required(Role.TEACHER)
def quiz_question_add(request, pk):
    quiz = get_object_or_404(_teacher_quizzes(request.user), pk=pk)
    form = QuestionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save(quiz)
        messages.success(request, "Savol qo‘shildi.")
    elif request.method == "POST":
        messages.error(request, form.errors.get("__all__", ["Xato."])[0])
    return redirect("portal:quiz_manage", pk=pk)


@role_required(Role.TEACHER)
def quiz_question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk, quiz__teacher_assignment__teacher=request.user)
    quiz_pk = question.quiz_id
    question.delete()
    messages.success(request, "Savol o‘chirildi.")
    return redirect("portal:quiz_manage", pk=quiz_pk)


@role_required(Role.TEACHER)
def quiz_toggle(request, pk):
    quiz = get_object_or_404(_teacher_quizzes(request.user), pk=pk)
    quiz.is_open = not quiz.is_open
    quiz.save(update_fields=["is_open"])
    return redirect("portal:quiz_manage", pk=pk)


@role_required(Role.TEACHER)
def quiz_results(request, pk):
    quiz = get_object_or_404(_teacher_quizzes(request.user), pk=pk)
    attempts = quiz.attempts.select_related("student").filter(submitted_at__isnull=False)
    ctx = {"panel_title": "O‘qituvchi paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "quiz": quiz, "attempts": attempts}
    return render(request, "portal/quiz/results.html", ctx)


# ==================== TALABA ====================

def _visible_quizzes(user):
    filters = Q()
    for e in StudentEnrollment.objects.filter(student=user):
        filters |= Q(
            teacher_assignment__department_course__course__curriculum__direction=e.direction,
            teacher_assignment__department_course__course__curriculum__study_form=e.study_form)
    if not filters:
        return Quiz.objects.none()
    return (Quiz.objects.filter(is_open=True).filter(filters)
            .select_related("teacher_assignment__department_course__course").distinct())


@role_required(Role.STUDENT)
def student_quizzes(request):
    quizzes = _visible_quizzes(request.user)
    done = {a.quiz_id: a for a in QuizAttempt.objects.filter(
        student=request.user, submitted_at__isnull=False)}
    rows = [{"quiz": q, "attempt": done.get(q.id)} for q in quizzes]
    ctx = {"panel_title": "Talaba paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "rows": rows}
    return render(request, "portal/quiz/student_list.html", ctx)


@role_required(Role.STUDENT)
def quiz_take(request, pk):
    quiz = get_object_or_404(_visible_quizzes(request.user), pk=pk)
    existing = QuizAttempt.objects.filter(quiz=quiz, student=request.user,
                                          submitted_at__isnull=False).first()
    if existing:
        return redirect("portal:quiz_result", pk=quiz.pk)

    questions = list(quiz.questions.prefetch_related("choices"))
    if request.method == "POST":
        attempt, _ = QuizAttempt.objects.get_or_create(quiz=quiz, student=request.user)
        attempt.answers.all().delete()
        for q in questions:
            ans = QuizAnswer.objects.create(attempt=attempt, question=q)
            if q.kind == Question.Kind.SHORT:
                ans.text_answer = request.POST.get(f"q{q.id}", "").strip()
                ans.save(update_fields=["text_answer"])
            else:
                ids = request.POST.getlist(f"q{q.id}")
                if ids:
                    ans.selected.set(Choice.objects.filter(id__in=ids, question=q))
        quiz_service.grade_attempt(attempt)
        messages.success(request, "Test topshirildi va avtomatik baholandi.")
        return redirect("portal:quiz_result", pk=quiz.pk)

    ctx = {"panel_title": "Talaba paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "quiz": quiz, "questions": questions}
    return render(request, "portal/quiz/take.html", ctx)


@role_required(Role.STUDENT)
def quiz_result(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    attempt = get_object_or_404(QuizAttempt, quiz=quiz, student=request.user,
                                submitted_at__isnull=False)
    answers = {a.question_id: a for a in attempt.answers.prefetch_related("selected")}
    rows = []
    for q in quiz.questions.prefetch_related("choices"):
        rows.append({"q": q, "ans": answers.get(q.id)})
    ctx = {"panel_title": "Talaba paneli",
           "panel_scope": request.user.full_name or request.user.username,
           "active": "quizzes", "quiz": quiz, "attempt": attempt, "rows": rows}
    return render(request, "portal/quiz/result.html", ctx)


# ==================== Test savollari Excel import ====================

@role_required(Role.TEACHER)
def quiz_import_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    wb = Workbook(); ws = wb.active; ws.title = "Savollar"
    headers = ["Savol", "To‘g‘ri javob", "Variant 2", "Variant 3", "Variant 4", "Ball"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3D5EE1")
    ws.append(["2+2 nechaga teng?", "4", "3", "5", "22", 1])
    ws.append(["O‘zbekiston poytaxti?", "Toshkent", "Samarqand", "Buxoro", "Namangan", 1])
    widths = [46, 20, 18, 18, 18, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = BytesIO(); wb.save(buf)
    resp = HttpResponse(buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="test_savollari_shablon.xlsx"'
    return resp


@role_required(Role.TEACHER)
def quiz_import(request, pk):
    from openpyxl import load_workbook
    quiz = get_object_or_404(_teacher_quizzes(request.user), pk=pk)
    if request.method != "POST" or not request.FILES.get("file"):
        return redirect("portal:quiz_manage", pk=pk)
    try:
        wb = load_workbook(request.FILES["file"], data_only=True)
    except Exception:
        messages.error(request, "Faylni o‘qib bo‘lmadi. Excel (.xlsx) yuklang.")
        return redirect("portal:quiz_manage", pk=pk)
    ws = wb.active
    created = 0
    skipped = 0
    order = quiz.questions.count()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        cells = list(row) + [None] * (6 - len(row))
        text = str(cells[0]).strip()
        correct = str(cells[1]).strip() if cells[1] is not None else ""
        opts = [str(cells[j]).strip() for j in (2, 3, 4) if cells[j] not in (None, "")]
        if not text or not correct:
            skipped += 1
            continue
        try:
            points = int(cells[5]) if cells[5] else 1
        except (ValueError, TypeError):
            points = 1
        order += 1
        q = Question.objects.create(quiz=quiz, text=text, kind=Question.Kind.SINGLE,
                                    points=points, order=order)
        # to'g'ri javob + variantlar (tartib aralashtiriladi)
        import random
        variants = [(correct, True)] + [(o, False) for o in opts]
        random.shuffle(variants)
        for oi, (t, ok) in enumerate(variants, start=1):
            Choice.objects.create(question=q, text=t, is_correct=ok, order=oi)
        created += 1
    if created:
        messages.success(request, f"{created} ta savol import qilindi." + (f" {skipped} ta satr o‘tkazib yuborildi." if skipped else ""))
    else:
        messages.error(request, "Savol topilmadi. Shablon ustunlarini tekshiring.")
    return redirect("portal:quiz_manage", pk=pk)
