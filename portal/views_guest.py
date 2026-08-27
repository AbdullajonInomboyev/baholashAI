"""O'qituvchi: tizimda yo'q talabalar uchun BIR MARTALIK test sessiyasi.

Oqim: shablon yuklab olinadi -> talaba ismlari yuklanadi -> har biriga login/parol
avtomatik beriladi -> talaba bir marta kirib testni yechadi -> o'qituvchi natijani
ko'radi -> xohlasa talabalarni o'chiradi, xohlasa keyingi safarga saqlaydi.
"""
from io import BytesIO

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.crypto import get_random_string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from accounts.models import PasswordStatus, Role, User
from assessment.models import GuestSession, GuestStudent, Quiz, QuizAttempt
from teaching.models import TeacherAssignment
from .access import role_required


def _panel(request):
    return {"panel_title": "O‘qituvchi paneli",
            "panel_scope": request.user.full_name or request.user.username}


def _teacher_quizzes(user):
    return (Quiz.objects.filter(teacher_assignment__teacher=user)
            .select_related("teacher_assignment__department_course__course")
            .order_by("-created_at"))


@role_required(Role.TEACHER)
def guest_sessions(request):
    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        quiz = Quiz.objects.filter(pk=request.POST.get("quiz"),
                                   teacher_assignment__teacher=request.user).first()
        if not title or not quiz:
            messages.error(request, "Sessiya nomi va test tanlanishi shart.")
        else:
            s = GuestSession.objects.create(teacher=request.user, quiz=quiz, title=title)
            messages.success(request, "Sessiya yaratildi. Endi talabalarni yuklang.")
            return redirect("portal:guest_detail", pk=s.pk)
    ctx = _panel(request)
    ctx.update({"active": "guest",
                "sessions": request.user.guest_sessions.select_related("quiz").all(),
                "quizzes": _teacher_quizzes(request.user)})
    return render(request, "portal/guest/sessions.html", ctx)


def _owned(request, pk):
    return get_object_or_404(GuestSession, pk=pk, teacher=request.user)


@role_required(Role.TEACHER)
def guest_detail(request, pk):
    s = _owned(request, pk)
    students = s.students.select_related("user")
    # natijalar (QuizAttempt orqali)
    attempts = {a.student_id: a for a in QuizAttempt.objects.filter(
        quiz=s.quiz, student__in=[st.user_id for st in students])}
    rows = []
    for st in students:
        att = attempts.get(st.user_id)
        rows.append({"st": st, "score": att.score if att and att.submitted_at else None,
                     "done": bool(att and att.submitted_at)})
    ctx = _panel(request)
    ctx.update({"active": "guest", "s": s, "rows": rows})
    return render(request, "portal/guest/detail.html", ctx)


@role_required(Role.TEACHER)
def guest_template(request):
    wb = Workbook(); ws = wb.active; ws.title = "Talabalar"
    c = ws.cell(1, 1, "F.I.Sh."); c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="274690")
    ws.column_dimensions["A"].width = 42
    for nm in ["Aliyev Vali", "Karimova Dilnoza", "Toshmatov Sardor"]:
        ws.append([nm])
    buf = BytesIO(); wb.save(buf)
    resp = HttpResponse(buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="mehmon_talabalar_shablon.xlsx"'
    return resp


@role_required(Role.TEACHER)
def guest_import(request, pk):
    s = _owned(request, pk)
    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Fayl tanlanmadi.")
        return redirect("portal:guest_detail", pk=pk)
    try:
        wb = load_workbook(f, data_only=True); ws = wb.active
    except Exception:
        messages.error(request, "Faylni o‘qib bo‘lmadi. Excel (.xlsx) bo‘lsin.")
        return redirect("portal:guest_detail", pk=pk)

    start = s.students.count()
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        full = str(row[0]).strip()
        start += 1
        login = f"g{s.pk}-{start:02d}"
        pw = get_random_string(5, "abcdefghijkmnpqrstuvwxyz23456789")
        u = User.objects.create(username=login, full_name=full,
                                is_guest=True, password_status=PasswordStatus.ACTIVE)
        u.set_password(pw); u.save()
        from accounts.models import UserRole
        UserRole.objects.get_or_create(user=u, role=Role.STUDENT)
        GuestStudent.objects.create(session=s, user=u, full_name=full,
                                    login=login, plain_password=pw)
        created += 1
    messages.success(request, f"{created} ta talaba yuklandi va login/parol berildi.")
    return redirect("portal:guest_detail", pk=pk)


@role_required(Role.TEACHER)
def guest_logins_export(request, pk):
    s = _owned(request, pk)
    wb = Workbook(); ws = wb.active; ws.title = "Loginlar"
    fill = PatternFill("solid", fgColor="274690")
    for i, h in enumerate(["F.I.Sh.", "Login", "Parol", "Holat"], 1):
        c = ws.cell(1, i, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
    ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12; ws.column_dimensions["D"].width = 14
    for st in s.students.all():
        ws.append([st.full_name, st.login, st.plain_password,
                   "Topshirgan" if st.taken_at else "Kutilmoqda"])
    buf = BytesIO(); wb.save(buf)
    resp = HttpResponse(buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="sessiya_{s.pk}_loginlar.xlsx"'
    return resp


@role_required(Role.TEACHER)
def guest_students_delete(request, pk):
    """Sessiyadagi barcha mehmon talabalarni (foydalanuvchilari bilan) o'chiradi, sessiya qoladi."""
    s = _owned(request, pk)
    user_ids = list(s.students.values_list("user_id", flat=True))
    s.students.all().delete()
    User.objects.filter(id__in=user_ids, is_guest=True).delete()
    messages.success(request, "Talabalar o‘chirildi. Sessiya saqlanib qoldi.")
    return redirect("portal:guest_detail", pk=pk)


@role_required(Role.TEACHER)
def guest_session_delete(request, pk):
    s = _owned(request, pk)
    user_ids = list(s.students.values_list("user_id", flat=True))
    s.delete()
    User.objects.filter(id__in=user_ids, is_guest=True).delete()
    messages.success(request, "Sessiya va talabalar o‘chirildi.")
    return redirect("portal:guest_sessions")


@role_required(Role.TEACHER)
def guest_toggle(request, pk):
    s = _owned(request, pk)
    s.is_open = not s.is_open
    s.save(update_fields=["is_open"])
    messages.success(request, "Sessiya holati o‘zgartirildi.")
    return redirect("portal:guest_detail", pk=pk)
