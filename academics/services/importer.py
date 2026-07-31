"""O‘quv reja (Excel) importi.

Shablon tuzilishi qat'iy: sarlavha 28–33-qatorlarda, ma'lumot 34-qatordan.
Faqat fanlar soni farq qiladi. Ustunlar birlashtirilgan, o‘rni doimiy.
Har bir varaq — bitta ta'lim shakli (kunduzgi/kechki/masofaviy).
"""
import re
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as col

from django.conf import settings
from django.db import transaction

from academics.models import (
    AcademicYear, Course, CourseSemester, Curriculum, Direction, ExcelImport, Semester, StudyForm,
)

DATA_START_ROW = 34
DIRECTION_CELL = ("S", 6)
REAL_CATEGORIES = {"Majburiy fanlar", "Tanlov fanlari"}

COLS = {
    "code": "B", "name": "C", "total": "V", "lecture": "AE", "practice": "AH",
    "lab": "AK", "seminar": "AN", "project": "AQ", "independent": "AT", "credit": "CS",
}
SEMESTER_HOURS = ["AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR"]   # 1..8
SEMESTER_CREDITS = ["BU", "BX", "CA", "CD", "CG", "CJ", "CM", "CP"]  # 1..8

SHEET_TO_FORM = [
    ("kund", StudyForm.FULL_TIME),
    ("kechki", StudyForm.EVENING),
    ("masofa", StudyForm.DISTANCE),
]


def _num(value):
    return value if isinstance(value, (int, float)) else None


def _form_for_sheet(title):
    lowered = title.lower()
    for marker, form in SHEET_TO_FORM:
        if marker in lowered:
            return form
    return None


def parse_sheet(ws):
    """Bitta varaqdan yo‘nalish va fanlar ro‘yxatini ajratib oladi."""
    direction = None
    raw = str(ws[f"{DIRECTION_CELL[0]}{DIRECTION_CELL[1]}"].value or "").strip()
    match = re.match(r"\s*(\d{6,8})\s*[-–]\s*(.+)", raw)
    if match:
        direction = {"code": match.group(1), "name": match.group(2).strip()}

    subjects = []
    category = None
    for row in range(DATA_START_ROW, ws.max_row + 1):
        tr = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=col(COLS["code"])).value
        name = ws.cell(row=row, column=col(COLS["name"])).value

        if code is None and name is not None and re.match(r"^\d+\.0+$", str(tr).strip()):
            category = str(name).strip()
            continue
        if code is None or name is None or category not in REAL_CATEGORIES:
            continue

        placements = []
        for index in range(8):
            hours = _num(ws.cell(row=row, column=col(SEMESTER_HOURS[index])).value)
            credits = _num(ws.cell(row=row, column=col(SEMESTER_CREDITS[index])).value)
            if hours or credits:
                placements.append({"semester": index + 1, "weekly_hours": hours, "credits": credits})

        subjects.append({
            "code": str(code).strip(),
            "name": str(name).strip(),
            "is_elective": category == "Tanlov fanlari",
            "total_hours": _num(ws.cell(row=row, column=col(COLS["total"])).value),
            "lecture_hours": _num(ws.cell(row=row, column=col(COLS["lecture"])).value),
            "practice_hours": _num(ws.cell(row=row, column=col(COLS["practice"])).value),
            "lab_hours": _num(ws.cell(row=row, column=col(COLS["lab"])).value),
            "seminar_hours": _num(ws.cell(row=row, column=col(COLS["seminar"])).value),
            "course_project_hours": _num(ws.cell(row=row, column=col(COLS["project"])).value),
            "independent_hours": _num(ws.cell(row=row, column=col(COLS["independent"])).value),
            "placements": placements,
        })
    return direction, subjects


def parse_workbook(path):
    wb = load_workbook(path, data_only=True)
    direction = None
    forms = {}
    for ws in wb.worksheets:
        form = _form_for_sheet(ws.title)
        if form is None:
            continue
        sheet_direction, subjects = parse_sheet(ws)
        direction = direction or sheet_direction
        forms[form] = subjects
    return direction, forms


@transaction.atomic
def import_curriculum(academic_year: AcademicYear, path, *, uploaded_by=None, forms=None, direction=None):
    """Excel'ni bazaga yozadi. `forms` — import qilinadigan ta'lim shakllari
    (bo‘sh bo‘lsa fayldagi barchasi). `direction` berilsa — o‘sha yo‘nalishga
    yoziladi (fayldagi kod o‘rniga), fanlar aralashmasligi uchun."""
    direction_data, parsed = parse_workbook(path)
    if direction is None:
        if not direction_data:
            raise ValueError("Yo‘nalish kodi topilmadi (S6 katak). Yo‘nalishni qo‘lda tanlang.")
        direction, _ = Direction.objects.get_or_create(
            code=direction_data["code"],
            defaults={"name": direction_data["name"], "faculty": academic_year.faculty},
        )

    selected = forms or list(parsed.keys())
    created_courses = 0
    for form in selected:
        subjects = parsed.get(form, [])
        curriculum, _ = Curriculum.objects.get_or_create(
            direction=direction, academic_year=academic_year, study_form=form
        )
        record = ExcelImport.objects.create(
            academic_year=academic_year, study_form=form,
            uploaded_by=uploaded_by, status=ExcelImport.Status.PROCESSING,
        )
        try:
            seen_codes = {}
            warnings = []
            for subject in subjects:
                code = subject["code"]
                if code in seen_codes:
                    warnings.append(
                        f"Takroriy kod: {code} — «{subject['name']}» va «{seen_codes[code]}». "
                        "Ikkinchisi o‘tkazib yuborildi, qo‘lda tuzatish kerak."
                    )
                    continue
                seen_codes[code] = subject["name"]
                course = Course.objects.update_or_create(
                    curriculum=curriculum, code=subject["code"],
                    defaults={
                        "name": subject["name"],
                        "is_elective": subject["is_elective"],
                        "total_hours": subject["total_hours"],
                        "lecture_hours": subject["lecture_hours"],
                        "practice_hours": subject["practice_hours"],
                        "lab_hours": subject["lab_hours"],
                        "seminar_hours": subject["seminar_hours"],
                        "course_project_hours": subject["course_project_hours"],
                        "independent_hours": subject["independent_hours"],
                        "credit": _credit(subject["total_hours"]),
                    },
                )[0]
                created_courses += 1
                for placement in subject["placements"]:
                    semester, _ = Semester.objects.get_or_create(
                        academic_year=academic_year, number=placement["semester"]
                    )
                    CourseSemester.objects.update_or_create(
                        course=course, semester=semester,
                        defaults={
                            "credits": _dec(placement["credits"]),
                            "weekly_hours": placement["weekly_hours"],
                        },
                    )
            imported = len(seen_codes)
            summary = f"{imported} ta fan import qilindi."
            if warnings:
                summary += "\n\nOgohlantirishlar:\n" + "\n".join(f"— {w}" for w in warnings)
            record.status = ExcelImport.Status.SUCCESS
            record.log = summary
            record.save(update_fields=["status", "log"])
        except Exception as exc:  # noqa: BLE001
            record.status = ExcelImport.Status.FAILED
            record.log = str(exc)
            record.save(update_fields=["status", "log"])
            raise

    return {"direction": direction, "courses": created_courses, "forms": selected}


def _credit(total_hours):
    if not total_hours:
        return None
    return round(Decimal(total_hours) / settings.ACADEMIC_HOURS_PER_CREDIT, 2)


def _dec(value):
    return Decimal(str(value)) if value is not None else None
